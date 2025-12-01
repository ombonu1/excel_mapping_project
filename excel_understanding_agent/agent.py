# understanding_agent.py
# """Understanding Agent for Excel process mapping.
# - Takes uploaded Excel files (Streamlit UploadedFile objects)
# - Builds a JSON metadata structure describing files, sheets, columns, sample rows
# - Sends that metadata to Gemini via google-genai
# - Returns a JSON dict describing mapping + inefficiencies

import os
import json
from io import BytesIO
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
import google.genai as genai
from google.genai.types import GenerateContentConfig

#1. Gemini client setup# ---------------------------------------------------------------------
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY")
if not GOOGLE_API_KEY: 
    raise RuntimeError("GOOGLE_API_KEY environment variable not set.")

from google.genai import Client
client = Client(api_key=GOOGLE_API_KEY)
model = "gemini-2.0-flash"


#2. Build metadata from uploaded Excel files#
def _build_files_metadata_from_uploaded(uploaded_files: List[Any]) -> Dict[str, Any]: 
    """ 
    Convert a list of uploaded Excel files into the 'files_metadata' JSON structure expected by the understanding agent. 
    """ 
    files_metadata: Dict[str, Any] = {"files": []} 

    for uploaded in uploaded_files: 
        uploaded.seek(0) 
        file_bytes = uploaded.read()
    
        def make_buffer() -> BytesIO: 
            return BytesIO(file_bytes) 
    
        try:
            wb = load_workbook(filename=make_buffer(), read_only=True, data_only=True) 
            sheet_names = wb.sheetnames 
        except Exception: 
            xl = pd.ExcelFile(make_buffer()) 
            sheet_names = xl.sheet_names 
        
        file_info = {"file_name": uploaded.name, "sheets": []} 
    
        for sheet_name in sheet_names: 
            try: 
                df = pd.read_excel(make_buffer(), sheet_name=sheet_name, engine="openpyxl") 
            except Exception: 
                continue 
        
            file_info["sheets"].append({ 
                "sheet_name": sheet_name, 
                "columns": [str(c) for c in df.columns], 
                "row_count": int(len(df)), 
                "sample_rows": df.head(5).to_dict(orient="records"), 
            }) 

        files_metadata["files"].append(file_info) 
    return files_metadata


## 3. Main agent wrapper for Streamlit # --------------
def run_understanding_agent(uploaded_files: List[Any]) -> Dict[str, Any]: 
    if not uploaded_files: 
	    raise ValueError("No files uploaded.") 

files_metadata = _build_files_metadata_from_uploaded(uploaded_files) 
metadata_json_str = json.dumps(files_metadata, indent=2) 

prompt = f"""
You are an expert agent for analysing collections of Excel spreadsheets used in business processes.
You are given a JSON object called `files_metadata` that describes the uploaded Excel files:

```json
{metadata_json_str}

Your tasks:

PART 1: MAP HOW THE SPREADSHEETS ARE USED
Read and understand the file structures.
Identify duplicated or overlapping data across sheets/files.
Track relationships between spreadsheets (shared IDs, lookup keys, reference columns).
Describe the role each spreadsheet plays in the process (input, transformation, lookup, export).
Describe the overall workflow: how data moves from one file to another.

PART 2: IDENTIFY INEFFICIENCIES & CONSOLIDATION OPPORTUNITIES
Spot repeated formulas or manual calculations.
Detect manually maintained copies of the same data.
Highlight spreadsheets that behave like pseudo-databases.
Identify frictions or risks in downstream workflows.
Call out governance/control risks (no validation, inconsistent coding, duplicated logic).

Respond ONLY with a single JSON object with these top-level keys:
"process_map"
"issues"
"opportunities"
"""
response = client.models.generate_content(
        model=model,
        contents=prompt,
        config=GenerateContentConfig(
            temperature=0.2,
            response_mime_type="application/json",
        ),
    )
    
result = json.loads(response.text)
return result
