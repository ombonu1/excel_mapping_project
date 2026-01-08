from typing import List, Dict, Any
from google.adk.tools import FunctionTool, ToolContext
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

# Import the shared bucket
from file_store import FILES

def build_files_metadata(files: List[str], tool_context: ToolContext) -> Dict[str, Any]:
    """
    Extracts metadata from Excel files stored in the global FILE_STORAGE.
    """
    print(f"\n=== TOOL CALLED: build_files_metadata ===")
    print(f"Requested files: {files}")

    # 1. Identify the Session
    # We need the session_id to find the right user's files in our bucket
    try:
        session_id = tool_context.session.session_id
    except AttributeError:
        # Fallback for local testing if context is mocked or missing
        print("Warning: No session_id found in context. Defaulting to 'default_session'")
        session_id = "default_session"

    print(f"Looking for files in session bucket: '{session_id}'")

    # 2. Retrieve the file map from the Sidecar Storage
    stored_files = FILES.get(session_id, {})
    
    if not stored_files:
        print(f"ERROR: No files found in FILES for session '{session_id}'")
        return {
            "files_metadata": {
                "files": [{"file_name": "unknown", "error": "No files uploaded to storage."}]
            }
        }

    # 3. Process the files
    result = {"files": []}
    
    for filename in files:
        # Check if we have the bytes for this specific file
        if filename not in stored_files:
            print(f"File '{filename}' not found in storage. Available: {list(stored_files.keys())}")
            result["files"].append({
                "file_name": filename, 
                "error": "File bytes not found in storage"
            })
            continue
            
        file_bytes = stored_files[filename]
        print(f"Processing '{filename}' ({len(file_bytes)} bytes)...")
        
        file_info = {"file_name": filename, "sheets": []}
        
        # Use BytesIO to work with the bytes (Standard Pandas Logic)
        excel_buffer = BytesIO(file_bytes)
        
        try:
            # Try openpyxl first (faster for just sheet names)
            wb = load_workbook(excel_buffer, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            # Fallback to pandas
            excel_buffer.seek(0)
            try:
                xl = pd.ExcelFile(excel_buffer)
                sheet_names = xl.sheet_names
            except Exception as e:
                file_info["error"] = f"Could not read Excel file: {str(e)}"
                result["files"].append(file_info)
                continue
        
        # Process each sheet
        for sheet_name in sheet_names:
            excel_buffer.seek(0)
            try:
                # Read just a few rows to get structure
                df = pd.read_excel(excel_buffer, sheet_name=sheet_name, nrows=5)
                
                # Convert timestamps to strings to avoid JSON serialization errors
                sample_data = df.head(3).astype(object).where(pd.notnull(df), None)
                
                sheet_info = {
                    "sheet_name": sheet_name,
                    "columns": list(map(str, df.columns)),
                    "column_count": len(df.columns),
                    "sample_rows": sample_data.to_dict("records") if len(df) > 0 else []
                }
                file_info["sheets"].append(sheet_info)
            except Exception as e:
                file_info["sheets"].append({
                    "sheet_name": sheet_name,
                    "error": f"Could not read sheet: {str(e)}"
                })
        
        result["files"].append(file_info)
    
    return {"files_metadata": result}

# Create the tool definition
files_metadata_tool = FunctionTool(
    func=build_files_metadata
)